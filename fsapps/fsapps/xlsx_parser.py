from collections import defaultdict
from pathlib import Path
import openpyxl

from helper import insert_data


def get_table_coordinate(worksheet):
    first_occurrence = {}
    last_occurrence = {}
    label_coords = {}

    for row in worksheet.iter_rows(
        min_row=2, min_col=worksheet.min_column, max_col=worksheet.max_column
    ):
        cell, label = row[0], row[1]
        value = cell.value

        if value not in first_occurrence:
            label_coords[value] = cell.row
            first_occurrence[value] = cell.row + 2
        last_occurrence[value] = cell.row

    return [first_occurrence, last_occurrence, label_coords]


def parse_excel(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook["DTS Report"]
    table_data = defaultdict(list)
    start_coords, end_coords, label_coords = get_table_coordinate(worksheet)

    for start, end, table in zip(
        start_coords.values(), end_coords.values(), label_coords.values()
    ):
        table_vals = next(
            worksheet.iter_rows(
                min_row=table,
                max_row=table,
                min_col=worksheet.min_column,
                max_col=worksheet.max_column,
            )
        )
        table_name = f"Table {table_vals[0].value} {table_vals[1].value}".strip()
        for row in worksheet.iter_rows(
            min_row=start,
            max_row=end,
            min_col=worksheet.min_column + 1,
            max_col=worksheet.max_column,
        ):
            row_vals = [i.value for i in row]
            label = row_vals.pop(0)
            row_vals = [val for val in row_vals if val is not None]
            if row_vals:
                table_data[table_name].append({label: row_vals})

    return table_data


def main():
    files_path = (Path.cwd().parent / "data").glob("*.xlsx")

    for item in files_path:
        print(f"Processing: {item.name}")
        result = parse_excel(item)

        if result:
            insert_data(result, item)


if __name__ == "__main__":
    main()
